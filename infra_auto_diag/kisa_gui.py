#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KISA 취약점 점검 GUI 도구 (원격 EC2/Linux SSH 실행기)
- 접속정보(IP/포트/계정/비밀번호 또는 SSH키)를 입력하고 [점검 실행]을 누르면
  같은 폴더의 kisa_check.py 를 원격 서버 /tmp 에 올려 실행하고,
  결과(JSON)를 회수해 표에 표시합니다.
- [엑셀로 추출] 버튼으로 .xlsx 저장.
- 읽기 전용 점검 스크립트를 실행하며, 접속정보는 화면 입력값만 사용합니다(저장/전송 안 함).

필요 패키지:
    pip install paramiko openpyxl
실행:
    python kisa_gui.py
"""
import os
import io
import re
import json
import zipfile
import datetime
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ---- 선택적 의존성 (없으면 안내) ----
try:
    import paramiko
except ImportError:
    paramiko = None
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_CHECK = os.path.join(SCRIPT_DIR, "kisa_unix_check.sh")  # 원격에 올릴 점검 셸 스크립트
REMOTE_SCRIPT = "/tmp/kisa_check.sh"
REMOTE_JSON = "/tmp/kisa_result.json"

# 결과보고서 양식(같은 폴더). 있으면 이 서식 그대로 채워서 저장한다.
REPORT_TEMPLATE = os.path.join(SCRIPT_DIR, "보고서_양식_Linux.xlsx")

# 양식 내부 시트 XML 경로(파일명 기준 — openpyxl 을 거치지 않고 직접 편집해야
# 그래프/도형이 보존된다).
#   sheet1 = 0. 표지            sheet2 = 2-1. 요약결과(그래프)_깨짐
#   sheet3 = 1. 진단 대상       sheet4 = 2-1. 요약결과(그래프)
#   sheet5 = 2-2. 요약 진단결과  sheet6 = 3-1. 진단 결과(Linux)
XL_COVER = "xl/worksheets/sheet1.xml"
XL_TARGET = "xl/worksheets/sheet3.xml"
XL_GRAPH = "xl/worksheets/sheet4.xml"
XL_SUMMARY = "xl/worksheets/sheet5.xml"
XL_DETAIL = "xl/worksheets/sheet6.xml"
XL_WORKBOOK = "xl/workbook.xml"

# 점검 스크립트 판정 → 보고서 판정 매핑
#   N/A(서비스/파일 미설치)  → 양호 로 기재(근거에 "해당 없음" 사유 유지)
#   수동확인(정책·주기 확인)  → "인터뷰 필요" 로 기재. 보안 적용율·영역별 점수 계산에서 제외한다.
MANUAL_LABEL = "인터뷰 필요"
# 점검 스크립트 판정 / 검증자 최종판정 → 보고서 기재값
#   수동확인(스크립트 용어) → "인터뷰 필요"(보고서 용어) 로 통일. 나머지는 그대로.
REPORT_STATUS = {
    "양호": "양호", "취약": "취약", "N/A": "N/A",
    "수동확인": MANUAL_LABEL, MANUAL_LABEL: MANUAL_LABEL,
}
# 검증자가 최종판정 대화상자에서 고를 수 있는 값
FINAL_CHOICES = ("양호", "취약", "N/A", "수동확인", MANUAL_LABEL)

# 보안 적용율(양호/진단항목) — 분모에서 N/A 와 "인터뷰 필요" 제외
_APPLY_RATE = (
    '(COUNTIF({c}$6:{c}$72,"양호"))/(COUNTA({c}$6:{c}$72)'
    '-COUNTIF({c}$6:{c}$72,"N/A")-COUNTIF({c}$6:{c}$72,"' + MANUAL_LABEL + '"))')


# ---------------- 양식 xlsx 직접 편집 헬퍼 ----------------
def _xesc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _cell_match(xml, ref):
    return re.search(r'<c r="' + re.escape(ref) + r'"([^>]*?)(?:/>|>.*?</c>)', xml, re.S)


def _cell_style(attrs):
    m = re.search(r'\bs="\d+"', attrs)
    return " " + m.group(0) if m else ""


def _cell_replace(xml, ref, builder, required=True):
    m = _cell_match(xml, ref)
    if not m:
        if required:
            raise KeyError("cell %s not found" % ref)
        return xml
    return xml[:m.start()] + builder(_cell_style(m.group(1))) + xml[m.end():]


def _put_str(xml, ref, text, required=True, s=None):
    """s 를 주면 셀 스타일 인덱스를 그 값으로 교체(배경색 등), 안 주면 원래 스타일 유지."""
    def build(old):
        st = f' s="{s}"' if s is not None else old
        return (f'<c r="{ref}"{st} t="inlineStr"><is><t xml:space="preserve">'
                f'{_xesc(text)}</t></is></c>')
    return _cell_replace(xml, ref, build, required)


def _put_num(xml, ref, num, required=True):
    return _cell_replace(xml, ref, lambda s: f'<c r="{ref}"{s}><v>{num}</v></c>', required)


def _put_formula(xml, ref, formula, required=True):
    esc = _xesc(formula)
    return _cell_replace(
        xml, ref, lambda s: f'<c r="{ref}"{s}><f>{esc}</f><v/></c>', required)


def _clear_cols(xml, row, col_from, col_to):
    """행 안에서 값/수식이 든 셀만 골라 빈 셀로 만든다(스타일 유지)."""
    for c in range(col_from, col_to + 1):
        ref = f"{_col_letter(c)}{row}"
        m = _cell_match(xml, ref)
        if not m:
            continue
        frag = m.group(0)
        if "<f>" in frag or "<v>" in frag or "<is>" in frag:
            xml = (xml[:m.start()]
                   + f'<c r="{ref}"{_cell_style(m.group(1))}/>'
                   + xml[m.end():])
    return xml


def _inject_verdict_fills(styles_xml):
    """styles.xml 의 <fills>·<cellXfs> 끝에 판정별 배경색 스타일을 추가한다.

    양식에는 '취약' 글자색을 빨강으로 바꾸는 조건부서식이 있으나 행 범위가
    들쭉날쭉해 일부 셀만 적용된다. 판정 셀에 배경색 스타일을 직접 지정해
    조건부서식과 무관하게 색이 일관되게 나오도록 한다.
    반환: (수정된 styles_xml, {판정값: cellXf 인덱스})
    """
    palette = {
        "취약": "FFF8D7DA", "양호": "FFD4EDDA",
        "N/A": "FFFFF3CD", MANUAL_LABEL: "FFD1ECF1",
    }
    mf = re.search(r'<fills count="(\d+)">', styles_xml)
    n_fills = int(mf.group(1))
    fills_add = "".join(
        f'<fill><patternFill patternType="solid"><fgColor rgb="{rgb}"/>'
        f'<bgColor rgb="FF000000"/></patternFill></fill>' for rgb in palette.values())
    styles_xml = (styles_xml[:mf.start()] + f'<fills count="{n_fills + len(palette)}">'
                  + styles_xml[mf.end():]).replace("</fills>", fills_add + "</fills>", 1)

    mx = re.search(r'<cellXfs count="(\d+)">', styles_xml)
    n_xfs = int(mx.group(1))
    xf_tpl = ('<xf numFmtId="0" fontId="10" fillId="{fid}" borderId="6" xfId="0" '
              'applyAlignment="1" applyFill="1">'
              '<alignment horizontal="center" vertical="center" wrapText="1"/></xf>')
    xfs_add = "".join(xf_tpl.format(fid=n_fills + i) for i in range(len(palette)))
    styles_xml = (styles_xml[:mx.start()] + f'<cellXfs count="{n_xfs + len(palette)}">'
                  + styles_xml[mx.end():]).replace("</cellXfs>", xfs_add + "</cellXfs>", 1)

    idx = {k: n_xfs + i for i, k in enumerate(palette)}
    return styles_xml, idx


STATUS_COLORS = {          # 표 행 배경색
    "취약": "#f8d7da",
    "양호": "#d4edda",
    "N/A": "#fff3cd",
    "수동확인": "#d1ecf1",
    "인터뷰 필요": "#d1ecf1",
}
XLSX_FILL = {              # 엑셀 셀 채우기 (ARGB 앞 2자리는 alpha)
    "취약": "F8D7DA",
    "양호": "D4EDDA",
    "N/A": "FFF3CD",
    "수동확인": "D1ECF1",
    "인터뷰 필요": "D1ECF1",
}

# 창 아이콘(logo.png) — 실행 위치와 무관하게 스크립트 폴더 기준, 없으면 조용히 무시
LOGO_PATH = os.path.join(SCRIPT_DIR, "logo.png")


def set_window_icon(win):
    try:
        img = tk.PhotoImage(file=LOGO_PATH)
        win._icon_ref = img          # GC 방지용 참조 유지
        win.iconphoto(False, img)
    except Exception:
        pass


class App:
    def __init__(self, root):
        self.root = root
        self.results = []          # 회수된 점검 결과 리스트
        self.host_label = ""
        self.os_label = ""
        self.conn_host = ""
        # 게이트웨이(bastion) 경유 설정
        self.gateway = {"enabled": False, "host": "", "port": 22, "user": "ec2-user",
                        "auth": "password", "password": "", "key": ""}
        root.title("취약점 빼기 팀 infra 진단 자동화 툴")
        root.geometry("1000x680")
        self._build_ui()

    # ---------------- UI 구성 ----------------
    def _build_ui(self):
        conn = ttk.LabelFrame(self.root, text="접속 정보")
        conn.pack(fill="x", padx=10, pady=8)

        # 1행: 호스트/포트/계정
        r1 = ttk.Frame(conn); r1.pack(fill="x", padx=8, pady=4)
        ttk.Label(r1, text="IP/호스트").grid(row=0, column=0, sticky="w")
        self.e_host = ttk.Entry(r1, width=24); self.e_host.grid(row=0, column=1, padx=6)
        ttk.Label(r1, text="포트").grid(row=0, column=2, sticky="w")
        self.e_port = ttk.Entry(r1, width=6); self.e_port.insert(0, "22"); self.e_port.grid(row=0, column=3, padx=6)
        ttk.Label(r1, text="계정(ID)").grid(row=0, column=4, sticky="w")
        self.e_user = ttk.Entry(r1, width=16); self.e_user.insert(0, "ec2-user"); self.e_user.grid(row=0, column=5, padx=6)

        # 2행: 인증 방식
        r2 = ttk.Frame(conn); r2.pack(fill="x", padx=8, pady=4)
        self.auth = tk.StringVar(value="password")
        ttk.Radiobutton(r2, text="비밀번호", variable=self.auth, value="password",
                        command=self._toggle_auth).grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(r2, text="SSH 키", variable=self.auth, value="key",
                        command=self._toggle_auth).grid(row=0, column=1, sticky="w", padx=(0, 12))

        ttk.Label(r2, text="비밀번호").grid(row=0, column=2, sticky="w")
        self.e_pass = ttk.Entry(r2, width=18, show="*"); self.e_pass.grid(row=0, column=3, padx=6)

        ttk.Label(r2, text="키 파일").grid(row=0, column=4, sticky="w")
        self.e_key = ttk.Entry(r2, width=26); self.e_key.grid(row=0, column=5, padx=6)
        self.btn_key = ttk.Button(r2, text="찾기", width=6, command=self._browse_key)
        self.btn_key.grid(row=0, column=6)
        self.btn_gw = ttk.Button(r2, text="게이트웨이 설정", command=self._open_gateway_dialog)
        self.btn_gw.grid(row=0, column=7, padx=(10, 0))

        # 3행: sudo 비밀번호 + 실행 버튼
        r3 = ttk.Frame(conn); r3.pack(fill="x", padx=8, pady=4)
        ttk.Label(r3, text="sudo 비밀번호(선택)").grid(row=0, column=0, sticky="w")
        self.e_sudo = ttk.Entry(r3, width=18, show="*"); self.e_sudo.grid(row=0, column=1, padx=6)

        self._toggle_auth()

        # 버튼들을 한 행에 묶어줄 프레임 생성
        self.btn_frame = ttk.Frame(self.root)
        self.btn_frame.pack(fill="x", padx=12) # 양옆 여백 설정

        # 상태 라벨
        self.lbl_sum = ttk.Label(self.btn_frame, text="대기 중…", anchor="w")
        self.lbl_sum.pack(side="left", padx=12, pady=5)


        # 엑셀 버튼 (오른쪽 정렬을 위해 오른쪽부터 채워 넣음)
        self.btn_xlsx = ttk.Button(self.btn_frame, text="엑셀로 추출", command=self.on_export, state="disabled")
        self.btn_xlsx.pack(side="right", padx=2)

        # 점검 실행 버튼 (엑셀 버튼 왼쪽에 배치됨)
        self.btn_run = ttk.Button(self.btn_frame, text="▶ 점검 실행", command=self.on_run)
        self.btn_run.pack(side="right", padx=2)

        # 결과 표
        table_frame = ttk.Frame(self.root)
        table_frame.pack(fill="both", expand=True, padx=10, pady=6)
        cols = ("code", "imp", "title", "auto", "final", "evidence")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings")
        for c, t, w in (("code", "코드", 60), ("imp", "중요도", 55), ("title", "점검 항목", 240),
                        ("auto", "자동판정", 75), ("final", "최종판정", 75), ("evidence", "근거", 430)):
            self.tree.heading(c, text=t)
            self.tree.column(c, width=w, anchor="w")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        for st, color in STATUS_COLORS.items():
            self.tree.tag_configure(st, background=color)
        self.tree.bind("<Double-1>", self._on_row_dblclick)
        ttk.Label(self.root,
                  text="↳ 행을 더블클릭 → 근거 확인 후 최종판정(양호/취약/N/A/인터뷰 필요) 지정. 엑셀은 '최종판정' 기준으로 추출되고 취약은 빨강으로 표시됩니다.",
                  foreground="#555").pack(fill="x", padx=12)

        # 로그
        logf = ttk.LabelFrame(self.root, text="로그")
        logf.pack(fill="x", padx=10, pady=6)
        self.txt = tk.Text(logf, height=7, wrap="word")
        self.txt.pack(fill="x", padx=4, pady=4)

        if paramiko is None:
            self.log("⚠ paramiko 미설치 → 터미널에서:  pip install paramiko")
        if openpyxl is None:
            self.log("⚠ openpyxl 미설치 → 터미널에서:  pip install openpyxl")
        if not os.path.exists(LOCAL_CHECK):
            self.log(f"⚠ 점검 스크립트 없음: {LOCAL_CHECK} (kisa_check.py를 같은 폴더에 두세요)")

    def _toggle_auth(self):
        if self.auth.get() == "password":
            self.e_pass.configure(state="normal")
            self.e_key.configure(state="disabled")
            self.btn_key.configure(state="disabled")
        else:
            self.e_pass.configure(state="disabled")
            self.e_key.configure(state="normal")
            self.btn_key.configure(state="normal")

    def _browse_key(self):
        path = filedialog.askopenfilename(title="SSH 개인키 선택")
        if path:
            self.e_key.delete(0, "end")
            self.e_key.insert(0, path)

    # ---------------- 게이트웨이(bastion) 설정 다이얼로그 ----------------
    def _open_gateway_dialog(self):
        g = self.gateway
        dlg = tk.Toplevel(self.root)
        set_window_icon(dlg)
        dlg.title("게이트웨이(Bastion) 설정")
        dlg.transient(self.root)
        dlg.resizable(False, False)
        dlg.grab_set()

        pad = dict(padx=6, pady=5)
        en = tk.BooleanVar(value=g["enabled"])
        ttk.Checkbutton(dlg, text="게이트웨이(bastion) 경유 접속 사용", variable=en)\
            .grid(row=0, column=0, columnspan=4, sticky="w", **pad)
        ttk.Label(dlg, text="private 서브넷 서버는 bastion을 점프호스트로 경유합니다.")\
            .grid(row=1, column=0, columnspan=4, sticky="w", padx=6)

        ttk.Label(dlg, text="Bastion IP").grid(row=2, column=0, sticky="e", **pad)
        e_host = ttk.Entry(dlg, width=24); e_host.insert(0, g["host"]); e_host.grid(row=2, column=1, **pad)
        ttk.Label(dlg, text="포트").grid(row=2, column=2, sticky="e", **pad)
        e_port = ttk.Entry(dlg, width=6); e_port.insert(0, str(g["port"])); e_port.grid(row=2, column=3, sticky="w", **pad)

        ttk.Label(dlg, text="계정(ID)").grid(row=3, column=0, sticky="e", **pad)
        e_user = ttk.Entry(dlg, width=24); e_user.insert(0, g["user"]); e_user.grid(row=3, column=1, **pad)

        auth = tk.StringVar(value=g["auth"])
        af = ttk.Frame(dlg); af.grid(row=4, column=0, columnspan=4, sticky="w", padx=6)
        ttk.Label(af, text="인증").grid(row=0, column=0)
        ttk.Radiobutton(af, text="비밀번호", variable=auth, value="password").grid(row=0, column=1, padx=6)
        ttk.Radiobutton(af, text="SSH 키", variable=auth, value="key").grid(row=0, column=2, padx=6)

        ttk.Label(dlg, text="비밀번호").grid(row=5, column=0, sticky="e", **pad)
        e_pass = ttk.Entry(dlg, width=24, show="*"); e_pass.insert(0, g["password"]); e_pass.grid(row=5, column=1, **pad)

        ttk.Label(dlg, text="키 파일").grid(row=6, column=0, sticky="e", **pad)
        e_key = ttk.Entry(dlg, width=30); e_key.insert(0, g["key"]); e_key.grid(row=6, column=1, columnspan=2, sticky="we", **pad)

        def browse():
            path = filedialog.askopenfilename(title="Bastion SSH 개인키 선택", parent=dlg)
            if path:
                e_key.delete(0, "end"); e_key.insert(0, path)
        ttk.Button(dlg, text="찾기", width=6, command=browse).grid(row=6, column=3, **pad)

        def save():
            self.gateway = {
                "enabled": en.get(),
                "host": e_host.get().strip(),
                "port": int(e_port.get().strip() or "22"),
                "user": e_user.get().strip() or "ec2-user",
                "auth": auth.get(),
                "password": e_pass.get(),
                "key": e_key.get().strip(),
            }
            self._refresh_gw_button()
            dlg.destroy()

        bf = ttk.Frame(dlg); bf.grid(row=7, column=0, columnspan=4, pady=8)
        ttk.Button(bf, text="저장", width=10, command=save).grid(row=0, column=0, padx=6)
        ttk.Button(bf, text="취소", width=10, command=dlg.destroy).grid(row=0, column=1, padx=6)

        dlg.wait_window()

    def _refresh_gw_button(self):
        g = self.gateway
        if g["enabled"] and g["host"]:
            self.btn_gw.configure(text=f"게이트웨이 ✓ ({g['host']})")
        else:
            self.btn_gw.configure(text="게이트웨이 설정")

    # ---------------- SSH 클라이언트 생성 (게이트웨이 지원) ----------------
    def _make_client(self, host, port, user, auth, password, key, sock=None):
        cli = paramiko.SSHClient()
        cli.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        kw = dict(hostname=host, port=port, username=user,
                  timeout=15, banner_timeout=15, auth_timeout=15)
        if sock is not None:
            kw["sock"] = sock
        if auth == "key":
            if not key or not os.path.exists(key):
                raise RuntimeError(f"키 파일 경로가 올바르지 않습니다: {key}")
            kw["key_filename"] = key
        else:
            kw["password"] = password
            kw["look_for_keys"] = False
            kw["allow_agent"] = False
        cli.connect(**kw)
        return cli

    # ---------------- 스레드 안전 UI 갱신 ----------------
    def log(self, msg):
        self.root.after(0, self._log, msg)

    def _log(self, msg):
        self.txt.insert("end", msg + "\n")
        self.txt.see("end")

    def set_busy(self, busy):
        self.root.after(0, lambda: self.btn_run.configure(state="disabled" if busy else "normal"))

    # ---------------- 실행 ----------------
    def on_run(self):
        if paramiko is None:
            messagebox.showerror("의존성 오류", "paramiko가 필요합니다.\npip install paramiko")
            return
        if not os.path.exists(LOCAL_CHECK):
            messagebox.showerror("파일 없음", f"점검 스크립트를 찾을 수 없습니다:\n{LOCAL_CHECK}")
            return
        host = self.e_host.get().strip()
        if not host:
            messagebox.showwarning("입력 필요", "IP/호스트를 입력하세요.")
            return
        params = dict(
            host=host,
            port=int(self.e_port.get().strip() or "22"),
            user=self.e_user.get().strip() or "ec2-user",
            auth=self.auth.get(),
            password=self.e_pass.get(),
            key=self.e_key.get().strip(),
            sudo=self.e_sudo.get(),
        )
        self.set_busy(True)
        self._clear_table()
        self.lbl_sum.configure(text="점검 실행 중…")
        threading.Thread(target=self._worker, args=(params,), daemon=True).start()

    def _worker(self, p):
        client = None
        gw_client = None
        try:
            g = self.gateway
            sock = None
            if g.get("enabled"):
                if not g.get("host"):
                    raise RuntimeError("게이트웨이 사용이 켜져 있으나 Bastion IP가 비어 있습니다.")
                self.log(f"[게이트웨이] {g['host']}:{g['port']} ({g['user']}) 접속 중…")
                gw_client = self._make_client(g["host"], g["port"], g["user"],
                                              g["auth"], g["password"], g["key"])
                self.log("      게이트웨이 접속 성공")
                tr = gw_client.get_transport()
                sock = tr.open_channel("direct-tcpip",
                                       (p["host"], p["port"]), ("127.0.0.1", 0))
                self.log(f"      터널 생성: bastion → {p['host']}:{p['port']}")

            via = " (게이트웨이 경유)" if sock else ""
            self.log(f"[1/4] {p['host']}:{p['port']} ({p['user']}) 접속 중…{via}")
            client = self._make_client(p["host"], p["port"], p["user"],
                                       p["auth"], p["password"], p["key"], sock=sock)
            self.log("      접속 성공")

            self.log("[2/4] 점검 스크립트 업로드…")
            sftp = client.open_sftp()
            sftp.put(LOCAL_CHECK, REMOTE_SCRIPT)

            self.log("[3/4] 원격 점검 실행 (수 초~수십 초 소요)…")
            base = f"bash {REMOTE_SCRIPT} --json {REMOTE_JSON} --no-color"
            if p["sudo"]:
                cmd = f"sudo -S -p '' {base}"
            else:
                cmd = f"sudo -n {base} 2>/dev/null || {base}"
            stdin, stdout, stderr = client.exec_command(cmd, get_pty=True, timeout=600)
            if p["sudo"]:
                stdin.write(p["sudo"] + "\n")
                stdin.flush()
            stdout.channel.recv_exit_status()   # 완료 대기

            self.log("[4/4] 결과 회수…")
            buf = io.BytesIO()
            sftp.getfo(REMOTE_JSON, buf)
            data = json.loads(buf.getvalue().decode("utf-8", "replace"))
            # 원격 임시파일 정리
            try:
                sftp.remove(REMOTE_JSON); sftp.remove(REMOTE_SCRIPT)
            except Exception:
                pass
            sftp.close()

            self.host_label = data.get("host", p["host"])
            self.os_label = data.get("os", "")
            self.conn_host = p["host"]
            self.results = data.get("results", [])
            self.log(f"      완료: {len(self.results)}개 항목 (OS: {data.get('os','?')})")
            self.root.after(0, self._show_results)
        except Exception as e:
            self.log(f"✖ 오류: {e}")
            self.root.after(0, lambda: messagebox.showerror("실행 오류", str(e)))
            self.root.after(0, lambda: self.lbl_sum.configure(text="오류 발생"))
        finally:
            if client:
                client.close()
            if gw_client:
                gw_client.close()
            self.set_busy(False)

    # ---------------- 결과 표시 ----------------
    def _clear_table(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

    def _show_results(self):
        self._clear_table()
        for idx, r in enumerate(self.results):
            r.setdefault("final", r.get("status", ""))   # 최종판정 초기값=자동판정
            r.setdefault("note", "")                      # 검증자 비고
            self._insert_row(idx, r)
        self._update_summary()
        self.btn_xlsx.configure(state="normal" if self.results else "disabled")

    def _insert_row(self, idx, r):
        fin = r.get("final", r.get("status", ""))
        self.tree.insert("", "end", iid=str(idx),
                         values=(r.get("code", ""), r.get("importance", ""),
                                 r.get("title", ""), r.get("status", ""), fin,
                                 " / ".join(r.get("evidence", []))),
                         tags=(fin,))

    def _update_row(self, idx):
        r = self.results[idx]
        fin = r.get("final", r.get("status", ""))
        self.tree.item(str(idx),
                       values=(r.get("code", ""), r.get("importance", ""),
                               r.get("title", ""), r.get("status", ""), fin,
                               " / ".join(r.get("evidence", []))),
                       tags=(fin,))

    def _update_summary(self):
        cnt = {k: 0 for k in ("양호", "취약", "N/A", "수동확인", MANUAL_LABEL)}
        for r in self.results:
            f = r.get("final", r.get("status", ""))
            cnt[f] = cnt.get(f, 0) + 1
        man = cnt["수동확인"] + cnt[MANUAL_LABEL]
        self.lbl_sum.configure(
            text=f"[{self.host_label}]  (최종판정 기준)  양호 {cnt['양호']}   취약 {cnt['취약']}   "
                 f"N/A {cnt['N/A']}   인터뷰필요 {man}   (총 {len(self.results)})")

    # ---------------- 행 더블클릭 → 상세/최종판정 ----------------
    def _on_row_dblclick(self, event):
        iid = self.tree.identify_row(event.y)
        if iid != "":
            self._open_detail(int(iid))

    def _open_detail(self, idx):
        r = self.results[idx]
        dlg = tk.Toplevel(self.root)
        set_window_icon(dlg)
        dlg.title(f"{r.get('code','')} 상세 / 최종판정")
        dlg.transient(self.root); dlg.grab_set(); dlg.geometry("580x480")

        head = ttk.Frame(dlg); head.pack(fill="x", padx=10, pady=8)
        ttk.Label(head, text=f"{r.get('code','')}  [{r.get('importance','')}]  {r.get('title','')}",
                  font=("", 11, "bold")).pack(anchor="w")
        ttk.Label(head, text=f"자동판정: {r.get('status','')}", foreground="#555").pack(anchor="w")

        ttk.Label(dlg, text="■ 근거", font=("", 9, "bold")).pack(anchor="w", padx=10)
        ev = tk.Text(dlg, height=9, wrap="word")
        ev.pack(fill="both", expand=True, padx=10, pady=(0, 6))
        ev.insert("1.0", "\n".join(r.get("evidence", []) or ["(근거 없음)"]))
        ev.configure(state="disabled")

        fr = ttk.LabelFrame(dlg, text="최종판정 (검증자 지정)")
        fr.pack(fill="x", padx=10, pady=4)
        fv = tk.StringVar(value=r.get("final", r.get("status", "")))
        for i, s in enumerate(FINAL_CHOICES):
            ttk.Radiobutton(fr, text=s, variable=fv, value=s).grid(
                row=i // 3, column=i % 3, padx=10, pady=4, sticky="w")

        ttk.Label(dlg, text="■ 비고(검증자 의견)", font=("", 9, "bold")).pack(anchor="w", padx=10)
        note = tk.Text(dlg, height=3, wrap="word")
        note.pack(fill="x", padx=10)
        note.insert("1.0", r.get("note", ""))

        def save():
            r["final"] = fv.get()
            r["note"] = note.get("1.0", "end").strip()
            self._update_row(idx)
            self._update_summary()
            self.log(f"  · {r.get('code','')} 최종판정 저장: {r['final']}")
            dlg.destroy()

        bf = ttk.Frame(dlg); bf.pack(pady=10)
        ttk.Button(bf, text="저장", width=10, command=save).grid(row=0, column=0, padx=8)
        ttk.Button(bf, text="닫기", width=10, command=dlg.destroy).grid(row=0, column=1, padx=8)
        dlg.wait_window()

    # ---------------- 엑셀 추출 ----------------
    def on_export(self):
        if openpyxl is None:
            messagebox.showerror("의존성 오류", "openpyxl이 필요합니다.\npip install openpyxl")
            return
        if not self.results:
            messagebox.showwarning("데이터 없음", "먼저 점검을 실행하세요.")
            return
        default = f"kisa_{self.host_label or 'result'}.xlsx".replace(":", "_")
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            initialfile=default,
                                            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        # 보고서 양식이 있으면 그 서식대로 채워서 저장
        if os.path.exists(REPORT_TEMPLATE):
            try:
                self._fill_template_raw(path)
                self.log(f"✔ 엑셀 저장(보고서 양식): {path}")
                messagebox.showinfo("저장 완료", f"보고서 양식으로 저장했습니다:\n{path}")
                return
            except PermissionError:
                self.log(f"✖ 저장 실패(권한 없음): {path}")
                messagebox.showerror(
                    "저장 실패",
                    "파일에 쓸 수 없습니다. 같은 이름의 파일이 Excel 등에서 "
                    f"열려 있으면 닫고 다시 저장하세요:\n{path}")
                return
            except Exception as e:
                self.log(f"⚠ 양식 채우기 실패({e}) → 기본 형식으로 저장합니다.")
        self._export_plain(path)

    def _fill_template_raw(self, path):
        """양식 xlsx 를 zip 단위로 직접 편집한다.

        openpyxl 로 열었다 저장하면 그래프(chart)·도형이 깨지므로, 시트 XML
        문자열만 손보고 charts/drawings/styles 등 나머지 파트는 원본 그대로
        다시 압축한다. 양식은 서버 15대용이므로 이번 점검(1대)에 안 쓰는 열은
        비워 #N/A·#DIV/0! 을 없앤다.
        """
        with zipfile.ZipFile(REPORT_TEMPLATE) as zin:
            order = zin.namelist()
            parts = {n: zin.read(n) for n in order}

        cover = parts[XL_COVER].decode("utf-8")
        target = parts[XL_TARGET].decode("utf-8")
        graph = parts[XL_GRAPH].decode("utf-8")
        summary = parts[XL_SUMMARY].decode("utf-8")
        detail = parts[XL_DETAIL].decode("utf-8")
        book = parts[XL_WORKBOOK].decode("utf-8")
        styles = parts["xl/styles.xml"].decode("utf-8")

        # 판정 셀 배경색 스타일 주입 (조건부서식이 일부 행만 색칠하는 문제 보완)
        styles, vstyle = _inject_verdict_fills(styles)

        host = (self.host_label or self.conn_host or "").strip()
        ipaddr = (self.conn_host or "").strip()
        osver = (self.os_label or "").strip()

        # 표지: 작성일 자동 기입
        cover = _put_str(cover, "B18", datetime.date.today().strftime("%Y. %m. %d."))

        # 진단 대상: 호스트 1대
        target = _put_str(target, "B1", "  ※ 진단 대상 리스트 - 서버 1대 (Linux 1대)")
        target = _put_num(target, "B5", 1)
        target = _put_str(target, "C5", host)
        target = _put_str(target, "D5", ipaddr)
        target = _put_str(target, "E5", osver)

        # 3-1 상세 결과: F=판정(최종판정 기준), G=근거
        by_code = {r.get("code", ""): r for r in self.results}
        for n in range(1, 68):                       # U-01 → 6행, U-67 → 72행
            r = by_code.get(f"U-{n:02d}")
            if not r:
                continue
            row = 5 + n
            raw = r.get("final") or r.get("status", "")        # 검증자 최종판정 우선
            verdict = REPORT_STATUS.get(raw, raw)
            detail = _put_str(detail, f"F{row}", verdict, s=vstyle.get(verdict))
            note = (r.get("note") or "").strip()
            evidence = " / ".join(r.get("evidence", []))
            if note:
                evidence = f"{evidence}  [검증자: {note}]" if evidence else f"[검증자: {note}]"
            detail = _put_str(detail, f"G{row}", evidence)
        for row in (3, 4, 5, 73, 74):               # 미사용 서버(H~AI) 정리
            detail = _clear_cols(detail, row, 8, 35)
        detail = _put_formula(detail, "F74", _APPLY_RATE.format(c="F"))

        # 2-2 요약: G~K열(서버2~6) 제거, "인터뷰 필요" 점수 제외
        summary = _put_formula(
            summary, "F72",
            "HLOOKUP(F$3,'3-1. 진단 결과(Linux)'!$F$3:$Q$72,ROW(A70),FALSE)")
        for row in range(3, 75):
            summary = _clear_cols(summary, row, 7, 11)
        for row in range(6, 73):
            rng = f"$F{row}:$T{row}"
            summary = _put_formula(
                summary, f"Z{row}",
                f'COUNTIF({rng},"N/A")+COUNTIF({rng},"{MANUAL_LABEL}")')
            summary = _put_formula(
                summary, f"W{row}",
                f'IF(COUNTIF({rng},"N/A")+COUNTIF({rng},"{MANUAL_LABEL}")'
                f'=COUNTA({rng}),"N/A",$X{row}/(COUNTA({rng})-$Z{row}))')
        summary = _put_formula(summary, "F74", _APPLY_RATE.format(c="F"))

        # 2-1 그래프: 서버 목록 1대로 축소 + 양식 버그(F53:T53 → F74:T74) 교정
        graph = _put_formula(
            graph, "D18",
            'COUNTIF(\'2-2. 요약 진단결과(Linux)\'!$F$74:$T$74,">=0.85")')
        graph = _put_formula(
            graph, "D20",
            'COUNTIF(\'2-2. 요약 진단결과(Linux)\'!$F$74:$T$74,"<0.7")')
        for row in range(37, 51):
            graph = _clear_cols(graph, row, 22, 24)   # V, W, X

        # workbook: 미완성 "_깨짐" 시트 숨김
        book = book.replace(
            '<sheet name="2-1. 요약결과(그래프)_깨짐" sheetId="2" state="visible" r:id="rId2" />',
            '<sheet name="2-1. 요약결과(그래프)_깨짐" sheetId="2" state="hidden" r:id="rId2" />')

        edited = {
            XL_COVER: cover, XL_TARGET: target, XL_GRAPH: graph,
            XL_SUMMARY: summary, XL_DETAIL: detail, XL_WORKBOOK: book,
        }
        import xml.dom.minidom as _minidom
        for name, xml in edited.items():
            _minidom.parseString(xml.encode("utf-8"))   # 형식 검증
            parts[name] = xml.encode("utf-8")
        # styles.xml 은 7MB대라 파싱이 느림 → 태그 균형만 가볍게 확인
        if styles.count("<fill>") == styles.count("</fill>") and styles.count("<cellXfs") == 1:
            parts["xl/styles.xml"] = styles.encode("utf-8")
        else:
            raise ValueError("styles.xml 편집 결과가 비정상")

        # 메모리에서 zip 을 완성한 뒤 마지막에 한 번만 기록한다
        # (기록 실패 시에도 반쯤 쓰인 손상 파일이 남지 않도록).
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in order:
                zout.writestr(name, parts[name])
        with open(path, "wb") as fh:
            fh.write(buf.getvalue())

    def _export_plain(self, path):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "KISA 점검결과"
            headers = ["호스트", "코드", "중요도", "점검 항목", "자동판정", "최종판정", "근거", "비고"]
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="404040")
                c.alignment = Alignment(horizontal="center")
            for r in self.results:
                raw = r.get("final") or r.get("status", "")
                fin = REPORT_STATUS.get(raw, raw)      # 수동확인 → 인터뷰 필요
                ws.append([self.host_label, r.get("code", ""), r.get("importance", ""),
                           r.get("title", ""), r.get("status", ""), fin,
                           " / ".join(r.get("evidence", [])), r.get("note", "")])
                fill = XLSX_FILL.get(fin)   # 최종판정 색상
                if fill:
                    for c in ws[ws.max_row]:
                        c.fill = PatternFill("solid", fgColor=fill)
            widths = [14, 7, 7, 30, 10, 10, 70, 24]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:H{ws.max_row}"
            wb.save(path)
            self.log(f"✔ 엑셀 저장: {path}")
            messagebox.showinfo("저장 완료", f"엑셀로 저장했습니다:\n{path}")
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))


def main():
    root = tk.Tk()
    set_window_icon(root)
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()