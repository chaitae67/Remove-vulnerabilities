#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""엑셀(보고서) 추출 모듈.

- export(app)            : 엑셀 버튼 진입점. 양식이 있으면 서식대로, 없으면 기본형식
- _fill_template_raw     : 보고서 양식(xlsx)을 zip 단위로 직접 편집(그래프 보존)
- _export_plain          : openpyxl 로 기본 표 형식 저장
그 외 _put_*, _clear_cols 등은 양식 셀을 직접 편집하는 헬퍼.
"""
import os
import io
import re
import zipfile
import datetime
from tkinter import filedialog, messagebox

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

from common import REPORT_STATUS, MANUAL_LABEL, SCRIPT_DIR

def _apply_rate(col, last):
    """보안 적용율(양호/진단항목) — 분모에서 N/A·인터뷰 필요 제외"""
    r = f"{col}$6:{col}${last}"
    return (f'(COUNTIF({r},"양호"))/(COUNTA({r})'
            f'-COUNTIF({r},"N/A")-COUNTIF({r},"{MANUAL_LABEL}"))')


# 계열별 보고서 양식 스펙 (openpyxl 을 거치지 않고 시트 XML 을 직접 편집 → 그래프 보존)
#   Linux/Windows 는 표지·깨짐 시트 순서만 다르다.
#   prefix/count/first_row/last_row : 3-1 상세결과의 항목 코드와 행 범위
#   score_range : 2-1 그래프가 참조하는 2-2 적용율 행 범위
#   summary_rewrite : 2-2 요약시트 처리 방식 (linux=서버열 직접정리 / windows=IFERROR 가드)
REPORT_SPECS = {
    "linux": {
        "template": os.path.join(SCRIPT_DIR, "보고서_양식_Linux.xlsx"),
        "cover": "xl/worksheets/sheet1.xml", "target": "xl/worksheets/sheet3.xml",
        "graph": "xl/worksheets/sheet4.xml", "summary": "xl/worksheets/sheet5.xml",
        "detail": "xl/worksheets/sheet6.xml",
        "broken_from": '<sheet name="2-1. 요약결과(그래프)_깨짐" sheetId="2" state="visible" r:id="rId2" />',
        "broken_to":   '<sheet name="2-1. 요약결과(그래프)_깨짐" sheetId="2" state="hidden" r:id="rId2" />',
        "label": "Linux", "prefix": "U", "count": 67, "first_row": 6, "last_row": 72,
        "detail_cols_from": 8, "detail_cols_to": 35,
        "score_range": "'2-2. 요약 진단결과(Linux)'!$F$74:$T$74",
        "summary_rewrite": "linux",
        "dxf_red": 4, "dxf_blue": 7,   # 이 양식 styles.xml 의 글꼴색 dxfId
    },
    "windows": {
        "template": os.path.join(SCRIPT_DIR, "보고서_양식_Windows.xlsx"),
        "cover": "xl/worksheets/sheet2.xml", "target": "xl/worksheets/sheet3.xml",
        "graph": "xl/worksheets/sheet4.xml", "summary": "xl/worksheets/sheet5.xml",
        "detail": "xl/worksheets/sheet6.xml",
        "broken_from": '<sheet name="2-1. 요약결과(그래프)_깨짐" sheetId="3" r:id="rId1"/>',
        "broken_to":   '<sheet name="2-1. 요약결과(그래프)_깨짐" sheetId="3" state="hidden" r:id="rId1"/>',
        "label": "Windows", "prefix": "W", "count": 64, "first_row": 6, "last_row": 69,
        "detail_cols_from": 8, "detail_cols_to": 13,
        "score_range": "'2-2. 요약 진단결과(Window)'!$F$71:$G$71",
        "summary_rewrite": "windows",
        "dxf_red": 22, "dxf_blue": 21,  # 이 양식 styles.xml 의 글꼴색 dxfId
    },
}


# ---------------- 양식 xlsx 직접 편집 헬퍼 ----------------
def _xesc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _cell_match(xml, ref):
    return re.search(r'<c r="' + re.escape(ref) + r'"([^>]*?)(?:/>|>.*?</c>)', xml, re.S)


def _cell_style(attrs):
    m = re.search(r'\bs="\d+"', attrs)
    return " " + m.group(0) if m else ""


def _cell_replace(xml, ref, builder, required=True):
    m = _cell_match(xml, ref)
    if not m:
        if required:
            raise KeyError("cell %s not found" % ref)
        return xml
    return xml[:m.start()] + builder(_cell_style(m.group(1))) + xml[m.end():]


def _put_str(xml, ref, text, required=True, s=None):
    """s 를 주면 셀 스타일 인덱스를 그 값으로 교체(배경색 등), 안 주면 원래 스타일 유지."""
    def build(old):
        st = f' s="{s}"' if s is not None else old
        return (f'<c r="{ref}"{st} t="inlineStr"><is><t xml:space="preserve">'
                f'{_xesc(text)}</t></is></c>')
    return _cell_replace(xml, ref, build, required)


def _put_num(xml, ref, num, required=True):
    return _cell_replace(xml, ref, lambda s: f'<c r="{ref}"{s}><v>{num}</v></c>', required)


def _put_formula(xml, ref, formula, required=True):
    esc = _xesc(formula)
    return _cell_replace(
        xml, ref, lambda s: f'<c r="{ref}"{s}><f>{esc}</f><v/></c>', required)


def _clear_cols(xml, row, col_from, col_to):
    """행 안에서 값/수식이 든 셀만 골라 빈 셀로 만든다(스타일 유지)."""
    for c in range(col_from, col_to + 1):
        ref = f"{_col_letter(c)}{row}"
        m = _cell_match(xml, ref)
        if not m:
            continue
        frag = m.group(0)
        if "<f>" in frag or "<v>" in frag or "<is>" in frag:
            xml = (xml[:m.start()]
                   + f'<c r="{ref}"{_cell_style(m.group(1))}/>'
                   + xml[m.end():])
    return xml


def _add_verdict_cf(sheet_xml, sqref, dxf_red, dxf_blue):
    """판정 셀 범위에 '취약=빨강 / 인터뷰 필요=파랑' 글꼴색 조건부서식을 추가한다.

    양식의 기존 조건부서식은 sqref 가 잘게 쪼개져 일부 행에만 '취약' 규칙이
    있어 색이 안 나온다. 우선순위 1~2 로 범위 전체를 덮는 규칙을 하나 더 넣어
    보완한다(양호는 규칙 없이 기본 검정 유지). styles.xml 은 건드리지 않으며,
    dxf_red/dxf_blue 는 계열별 양식 styles.xml 에 이미 있는 글꼴색 dxfId 다.
    """
    top = sqref.split(":")[0]           # 조건부서식 수식 기준 셀 (예: F6)
    block = (
        f'<conditionalFormatting sqref="{sqref}">'
        f'<cfRule type="containsText" dxfId="{dxf_red}" priority="1" operator="containsText" '
        f'text="취약"><formula>NOT(ISERROR(SEARCH("취약",{top})))</formula></cfRule>'
        f'<cfRule type="containsText" dxfId="{dxf_blue}" priority="2" operator="containsText" '
        f'text="인터뷰"><formula>NOT(ISERROR(SEARCH("인터뷰",{top})))</formula></cfRule>'
        f'</conditionalFormatting>')
    return sheet_xml.replace("<pageMargins", block + "<pageMargins", 1)


# ---------------- 클라우드(AWS/Azure/GCP) 양식 ----------------
#   서버 보고서와 달리 차트가 없는 단일 시트라 openpyxl 로 직접 채운다.
CLOUD_SPECS = {
    "AWS": {"template": os.path.join(SCRIPT_DIR, "보고서_양식_AWS.xlsx"),
            "sheet": "진단 결과(AWS)", "first_row": 4, "code_col": 2,
            "result_col": 7, "detail_col": 8, "resource_col": 9},
    "AZURE": {"template": os.path.join(SCRIPT_DIR, "보고서_양식_Azure.xlsx"),
              "sheet": "진단 결과(Azure)", "first_row": 4, "code_col": 2,
              "result_col": 7, "detail_col": 8, "resource_col": 9},
    # GCP 는 3단계에서 추가
}


def _cloud_spec(app):
    label = (getattr(app, "os_label", "") or "").strip().upper()
    return CLOUD_SPECS.get(label)


# ---------------- 진입점 ----------------
def _pick_spec(app):
    fam = getattr(app, "family", "") or app.os_choice.get()
    return REPORT_SPECS.get(fam, REPORT_SPECS["linux"])


def _fill_cloud_template(app, path, cspec):
    """클라우드 진단 결과를 단일 시트 양식(G/H/I 열)에 채운다(차트 없음 → openpyxl 직접)."""
    wb = openpyxl.load_workbook(cspec["template"])
    ws = wb[cspec["sheet"]]
    by_code = {str(r.get("code", "")): r for r in app.results}
    row_of = {}
    for row in range(cspec["first_row"], ws.max_row + 1):
        code = ws.cell(row=row, column=cspec["code_col"]).value
        if code is not None:
            row_of[str(code).strip()] = row

    red = Font(color="FFFF0000", bold=True)
    blue = Font(color="FF0070C0", bold=True)
    black = Font(color="FF000000")
    for code, r in by_code.items():
        row = row_of.get(code)
        if not row:
            continue
        raw = r.get("final") or r.get("status", "")
        verdict = REPORT_STATUS.get(raw, raw)
        note = (r.get("note") or "").strip()
        detail = " / ".join(r.get("evidence", []))
        if note:
            detail = f"{detail}  [검증자: {note}]" if detail else f"[검증자: {note}]"
        gc = ws.cell(row=row, column=cspec["result_col"], value=verdict)
        gc.font = red if verdict == "취약" else blue if verdict == MANUAL_LABEL else black
        ws.cell(row=row, column=cspec["detail_col"], value=detail)
        ws.cell(row=row, column=cspec["resource_col"],
                value=" / ".join(r.get("resources", []))[:32000])
    wb.save(path)


def export(app):
    """엑셀 버튼 진입점. 계열별 양식이 있으면 서식대로, 없으면 기본형식."""
    if openpyxl is None:
        messagebox.showerror("의존성 오류", "openpyxl이 필요합니다.\npip install openpyxl")
        return
    if not app.results:
        messagebox.showwarning("데이터 없음", "먼저 점검을 실행하세요.")
        return

    # 클라우드 진단 결과
    if (getattr(app, "family", "") or "").lower() == "cloud":
        cspec = _cloud_spec(app)
        default = f"클라우드_{app.os_label}_{app.host_label or 'result'}.xlsx".replace(":", "_")
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default,
                                            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        if cspec and os.path.exists(cspec["template"]):
            try:
                _fill_cloud_template(app, path, cspec)
                app.log(f"✔ 엑셀 저장({app.os_label} 클라우드 양식): {path}")
                messagebox.showinfo("저장 완료", f"{app.os_label} 진단 양식으로 저장했습니다:\n{path}")
                return
            except PermissionError:
                messagebox.showerror("저장 실패", f"파일이 열려 있으면 닫고 다시 저장하세요:\n{path}")
                return
            except Exception as e:
                app.log(f"⚠ 클라우드 양식 채우기 실패({e}) → 기본 형식으로 저장합니다.")
        _export_plain(app, path)
        return

    spec = _pick_spec(app)
    default = f"kisa_{spec['label']}_{app.host_label or 'result'}.xlsx".replace(":", "_")
    path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                        initialfile=default,
                                        filetypes=[("Excel", "*.xlsx")])
    if not path:
        return
    # 계열별 보고서 양식이 있으면 그 서식대로 채워서 저장
    if os.path.exists(spec["template"]):
        try:
            _fill_template_raw(app, path, spec)
            app.log(f"✔ 엑셀 저장({spec['label']} 보고서 양식): {path}")
            messagebox.showinfo("저장 완료", f"{spec['label']} 보고서 양식으로 저장했습니다:\n{path}")
            return
        except PermissionError:
            app.log(f"✖ 저장 실패(권한 없음): {path}")
            messagebox.showerror(
                "저장 실패",
                "파일에 쓸 수 없습니다. 같은 이름의 파일이 Excel 등에서 "
                f"열려 있으면 닫고 다시 저장하세요:\n{path}")
            return
        except Exception as e:
            app.log(f"⚠ 양식 채우기 실패({e}) → 기본 형식으로 저장합니다.")
    _export_plain(app, path)


def _fill_template_raw(app, path, spec):
    """계열별 보고서 양식 xlsx 를 zip 단위로 직접 편집한다.

    openpyxl 로 열었다 저장하면 그래프(chart)·도형이 깨지므로 시트 XML
    문자열만 손보고 charts/drawings/styles 등은 원본 그대로 다시 압축한다.
    양식은 서버 2대(또는 15대)용이라 이번 점검(1대)에 안 쓰는 열은 비운다.
    """
    first, last = spec["first_row"], spec["last_row"]
    pre = spec["prefix"]
    cf_range = f"F{first}:F{last}"

    with zipfile.ZipFile(spec["template"]) as zin:
        order = zin.namelist()
        parts = {n: zin.read(n) for n in order}

    cover = parts[spec["cover"]].decode("utf-8")
    target = parts[spec["target"]].decode("utf-8")
    graph = parts[spec["graph"]].decode("utf-8")
    summary = parts[spec["summary"]].decode("utf-8")
    detail = parts[spec["detail"]].decode("utf-8")
    book = parts["xl/workbook.xml"].decode("utf-8")
    # styles.xml 은 건드리지 않는다 — 색은 조건부서식(글꼴색)으로 처리

    host = (app.host_label or app.conn_host or "").strip()
    ipaddr = (app.conn_host or "").strip()
    osver = (app.os_label or "").strip()

    # 표지: 작성일 자동 기입
    cover = _put_str(cover, "B18", datetime.date.today().strftime("%Y. %m. %d."),
                     required=False)

    # 진단 대상: 호스트 1대
    target = _put_str(target, "B1",
                      f"  ※ 진단 대상 리스트 - 서버 1대 ({spec['label']} 1대)")
    target = _put_num(target, "B5", 1)
    target = _put_str(target, "C5", host)
    target = _put_str(target, "D5", ipaddr)
    target = _put_str(target, "E5", osver)
    target = _put_str(target, "F5", "-", required=False)   # 용도(미수집) — 0 표시 방지
    for row in range(6, 20):                     # 서버 2대 이상 슬롯 비우기
        target = _clear_cols(target, row, 2, 7)

    # 3-1 상세 결과: F=판정(최종판정 기준), G=근거
    by_code = {r.get("code", ""): r for r in app.results}
    for n in range(1, spec["count"] + 1):
        r = by_code.get(f"{pre}-{n:02d}")
        if not r:
            continue
        row = first - 1 + n
        raw = r.get("final") or r.get("status", "")        # 검증자 최종판정 우선
        verdict = REPORT_STATUS.get(raw, raw)
        detail = _put_str(detail, f"F{row}", verdict)
        note = (r.get("note") or "").strip()
        evidence = " / ".join(r.get("evidence", []))
        if note:
            evidence = f"{evidence}  [검증자: {note}]" if evidence else f"[검증자: {note}]"
        detail = _put_str(detail, f"G{row}", evidence)
    dc_from, dc_to = spec["detail_cols_from"], spec["detail_cols_to"]
    for row in (3, 4, 5, last + 1, last + 2):    # 미사용 서버 열(H~) 정리
        detail = _clear_cols(detail, row, dc_from, dc_to)
    detail = _put_formula(detail, f"F{last + 2}", _apply_rate("F", last))
    detail = _add_verdict_cf(detail, cf_range,   # 취약=빨강 / 인터뷰=파랑 글꼴색
                             spec["dxf_red"], spec["dxf_blue"])

    # 2-2 요약
    if spec["summary_rewrite"] == "linux":
        summary = _put_formula(
            summary, "F72",
            "HLOOKUP(F$3,'3-1. 진단 결과(Linux)'!$F$3:$Q$72,ROW(A70),FALSE)")
        for row in range(3, 75):
            summary = _clear_cols(summary, row, 7, 11)
        for row in range(6, 73):
            rng = f"$F{row}:$T{row}"
            summary = _put_formula(
                summary, f"Z{row}",
                f'COUNTIF({rng},"N/A")+COUNTIF({rng},"{MANUAL_LABEL}")')
            summary = _put_formula(
                summary, f"W{row}",
                f'IF(COUNTIF({rng},"N/A")+COUNTIF({rng},"{MANUAL_LABEL}")'
                f'=COUNTA({rng}),"N/A",$X{row}/(COUNTA({rng})-$Z{row}))')
        summary = _put_formula(
            summary, "V39",
            'IF(COUNTIF(W39:W68,"N/A")=COUNTA(W39:W68),"N/A",AVERAGE(W39:W68))')
    else:  # windows — 서버 2 열(G) 전체를 비운다(비우지 않으면 G71 이 0 으로
           # 계산돼 그래프 점수 범위 F71:G71 에 유령 '취약' 서버가 생긴다)
        for row in range(3, last + 3):          # G3(번호) ~ G71(보안 적용율)
            summary = _clear_cols(summary, row, 7, 7)
    summary = _put_formula(summary, f"F{last + 2}", _apply_rate("F", last))
    summary = _add_verdict_cf(summary, cf_range, spec["dxf_red"], spec["dxf_blue"])

    # 2-1 그래프 : SCORE = 2-2 의 서버별 보안 적용율 행
    #   안전 D18 = COUNTIF(SCORE,">=0.85") / 양호 D19 = COUNTIFS(SCORE,">=0.7",…,"<0.85")
    #   취약 D20 = COUNTIF(SCORE,"<0.7") / C5·C6 = AVERAGE(SCORE)
    SCORE = spec["score_range"]
    graph = _put_formula(graph, "D18", f'COUNTIF({SCORE},">=0.85")', required=False)
    graph = _put_formula(graph, "D19", f'COUNTIFS({SCORE},">=0.7",{SCORE},"<0.85")', required=False)
    graph = _put_formula(graph, "D20", f'COUNTIF({SCORE},"<0.7")', required=False)
    graph = _put_formula(graph, "C5", f'AVERAGE({SCORE})', required=False)
    graph = _put_formula(graph, "C6", f'AVERAGE({SCORE})', required=False)
    for row in range(37, 51):
        graph = _clear_cols(graph, row, 22, 24)   # 미사용 서버 목록(V/W/X)

    # workbook: 미완성 "_깨짐" 시트 숨김
    book = book.replace(spec["broken_from"], spec["broken_to"])

    # 열 때 수식을 강제로 재계산시킨다. 이 설정이 없으면(윈도우 양식이 그랬다)
    # 엑셀이 캐시된 옛 값/빈 값을 그대로 보여줘 보안 적용율·그래프가 안 나온다.
    if "<calcPr" not in book:
        book = book.replace(
            "</workbook>", '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')
    elif "fullCalcOnLoad" not in book:
        book = re.sub(r"<calcPr ", '<calcPr fullCalcOnLoad="1" ', book, count=1)

    edited = {
        spec["cover"]: cover, spec["target"]: target, spec["graph"]: graph,
        spec["summary"]: summary, spec["detail"]: detail, "xl/workbook.xml": book,
    }
    import xml.dom.minidom as _minidom
    for name, xml in edited.items():
        _minidom.parseString(xml.encode("utf-8"))   # 형식 검증
        parts[name] = xml.encode("utf-8")

    # 메모리에서 zip 을 완성한 뒤 마지막에 한 번만 기록한다
    # (기록 실패 시에도 반쯤 쓰인 손상 파일이 남지 않도록).
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in order:
            zout.writestr(name, parts[name])
    with open(path, "wb") as fh:
        fh.write(buf.getvalue())


def _export_plain(app, path):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "취약점 빼기 팀 자동화 결과"
        headers = ["호스트", "코드", "중요도", "점검 항목", "자동판정", "최종판정", "근거", "비고"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="404040")
            c.alignment = Alignment(horizontal="center")
        font_color = {"취약": "FFFF0000", "인터뷰 필요": "FF0070C0"}  # 양호=검정
        for r in app.results:
            raw = r.get("final") or r.get("status", "")
            fin = REPORT_STATUS.get(raw, raw)      # 수동확인 → 인터뷰 필요
            ws.append([app.host_label, r.get("code", ""), r.get("importance", ""),
                       r.get("title", ""), r.get("status", ""), fin,
                       " / ".join(r.get("evidence", [])), r.get("note", "")])
            fc = font_color.get(fin)
            if fc:
                ws.cell(row=ws.max_row, column=6).font = Font(bold=True, color=fc)
        widths = [14, 7, 7, 30, 10, 10, 70, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{ws.max_row}"
        wb.save(path)
        app.log(f"✔ 엑셀 저장: {path}")
        messagebox.showinfo("저장 완료", f"엑셀로 저장했습니다:\n{path}")
    except Exception as e:
        messagebox.showerror("저장 오류", str(e))
